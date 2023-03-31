import tkinter.messagebox
from tkinter import *
from tkinter import ttk, filedialog
from openpyxl import load_workbook
from tkinter.messagebox import askyesno

from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg,
                                               NavigationToolbar2Tk)


# plot function is created for
# plotting the graph in
# tkinter window
class App:

    def __init__(self, master):
        self.master = master
        self.STATE = {'click': 0}
        self.selectedIndex = None

        self.plot1 = None
        self.plot2 = None
        self.secondWindow = None
        self.canvas = None
        self.modifyDataBySelection = [0, 0, 0]  # [Drag and move, Delete before, Delete after]

        self.graphYMountylyOil = []
        self.graphYMountylyGas = []
        self.graphYMountylyWater = []
        self.graphYDailyOil = []
        self.graphYDailyGas = []
        self.graphYDailyWater = []
        self.graphYValues = []
        self.graphXValues = []
        self.secondaryGraphValues = []
        self.updatedGraphValues = []
        self.showSecondaryChart = IntVar()
        self.charts = ["Montly Oil", "Montly Gas", "Montly Water", "Daily Oil", "Daily Gas", "Daily Water"]
        self.chartNameColors = ['red', 'green', 'purple', 'orange', 'blue', 'brown']
        self.chartOnOffValue = [[3, False], [4, False], [5, False], [7, False], [8, False], [9, False]]
        self.selectedColumn = 0
        self.titles = []

        self.sheetNames = []
        self.saveType = IntVar()  # 0 for typed sheet save, 1 for save by sheet selection
        self.selectedSheetToSave = None
        self.typedSheetToSave = None

        # *** File Menu ***
        menu = Menu(master)  # Create a menu and put it in main window
        root.config(menu=menu)  # We are configuring a menu for this project called menu

        fileMenu = Menu(menu, tearoff=0)
        menu.add_cascade(label="File", menu=fileMenu)  # Add dropdown menu item
        fileMenu.add_command(label="Open Excel", command=self.open_file)  # Add item to dropdown menu item

        # *** Tab Selection Area ***
        self.tabControl = ttk.Notebook(root)
        self.tabControl.bind("<<NotebookTabChanged>>", self.selectTabMethod)

        self.tabs = dict()
        self.tabs['PAGE 1'] = ttk.Frame(self.tabControl)
        self.tabControl.add(self.tabs['PAGE 1'], text='Tab 1')
        self.tabControl.pack(expand=1, fill="both")
        self.selectedTab = self.tabControl.tab(self.tabControl.select(), "text")

        # *** Inside Tab Area ***
        self.toolbar = Frame(self.master, bg="lightgray")
        self.selectExcelFileButton = Button(self.toolbar, text="Select Excel File", command=self.doNothing)  # Check here later

        # *** Main Frame ***
        self.tabFrame = Frame(self.master)

        # *** Right and Left Frame ***
        self.canvasArea = Frame(self.tabFrame)

        rightFrame = Frame(self.canvasArea, bg="lightgray", width=150)
        leftFrame = Frame(self.canvasArea, bg="lightgray", width=150)

        self.chartListLabel = Label(rightFrame, text="Charts")
        self.pointListLabel = Label(rightFrame, text="Points")

        self.chartListboxFrame = Frame(rightFrame)
        self.pointListboxFrame = Frame(rightFrame)

        self.chartListbox = Listbox(self.chartListboxFrame)
        self.pointListbox = Listbox(self.pointListboxFrame)

        chartListboxScrollbar = Scrollbar(self.chartListboxFrame)
        chartListboxScrollbar.pack(side=RIGHT, fill=BOTH)
        self.chartListbox.config(yscrollcommand=chartListboxScrollbar.set)
        chartListboxScrollbar.config(command=self.chartListbox.yview)

        pointListboxScrollbar = Scrollbar(self.pointListboxFrame)
        pointListboxScrollbar.pack(side=RIGHT, fill=BOTH)
        self.pointListbox.config(yscrollcommand=pointListboxScrollbar.set)
        pointListboxScrollbar.config(command=self.pointListbox.yview)

        self.addRemoveChartButton = Button(rightFrame, text="Add/Remove Graph", command=self.addRemoveChart)
        self.showSecondaryChartCheckBox = Checkbutton(rightFrame, text='Show number of days that the well sas active', variable=self.showSecondaryChart,
                                       onvalue=1, offvalue=0, command=self.secondaryChartView)

        self.chartListbox.bind('<Double-Button>', lambda x: self.selectChart(self.chartListbox.curselection()[0]))
        # *** Fill in Right Frame ***
        self.chartListLabel.pack()
        self.chartListbox.pack()
        self.chartListboxFrame.pack()
        self.pointListLabel.pack()
        self.pointListbox.pack()
        self.pointListboxFrame.pack()
        self.addRemoveChartButton.pack(padx=2, pady=2)
        self.showSecondaryChartCheckBox.pack(padx=2, pady=2)
        rightFrame.pack(side=RIGHT, fill=Y)

        # *** Fill in Left Frame ***
        self.updateSelectionsButton = Button(leftFrame, text="Update Selections", state="disabled", command=self.updateSelectionMethod)
        self.moveYAxisButton = Button(leftFrame, text="Move Y Point", state="disabled", command=lambda: self.modifyDataBySelectionMethod(0))
        self.deleteDataBeforeButton = Button(leftFrame, text="Delete Data Before", state="disabled", command=lambda: self.modifyDataBySelectionMethod(1))
        self.deleteDataAfterButton = Button(leftFrame, text="Delete Data After", state="disabled", command=lambda: self.modifyDataBySelectionMethod(2))
        self.saveDataButton = Button(leftFrame, text="Save data", state="disabled", command=self.saveDataMethod)

        self.updateSelectionsButton.pack()
        self.moveYAxisButton.pack()
        self.deleteDataBeforeButton.pack()
        self.deleteDataAfterButton.pack()
        self.saveDataButton.pack()
        leftFrame.pack(side=LEFT, fill=Y)

        self.canvasArea.pack(fill=BOTH, expand=Y)

        # Create a Treeview widget
        self.tree = ttk.Treeview(self.canvasArea)

        self.master.bind("<Escape>", self.cancelModifyData)

    def selectTabMethod(self, event):
        id = self.tabControl.select()
        if self.tabControl.tab(id, "text") == "Tab 1":
            self.toolbar.pack(side=TOP, fill=X, in_=self.tabs['PAGE 1'])
            self.tabFrame.pack(fill=BOTH, expand=1, in_=self.tabs['PAGE 1'])

    def modifyDataBySelectionMethod(self, selectedIndex):
        for index, value in enumerate(self.modifyDataBySelection):
            if index == selectedIndex:
                self.modifyDataBySelection[index] = 1 - value
            else:
                self.modifyDataBySelection[index] = 0

    def cancelModifyData(self):
        self.modifyDataBySelection = [0, 0, 0]

    def createGraph(self):
        if not self.canvas:
            # the figure that will contain the plot
            fig = Figure(figsize=(7, 7),
                         dpi=100)
            # adding the subplot
            self.plot1 = fig.add_subplot(111)
        for index, column in enumerate(self.chartOnOffValue):
            if column[1] == True:
                y = []
                for value in self.graphYValues:
                    y.append(value[column[0]])
                # plotting the graph
                self.plot1.plot(self.graphXValues, y, marker='o', color=self.chartNameColors[index])
        if self.showSecondaryChart.get() == 1:
            self.plot2 = self.plot1.twinx()
            y = []
            for value in self.graphYValues:
                y.append(value[6])
            self.plot2.plot(self.graphXValues, y, 'ro', color='black')
        # creating the Tkinter canvas
        # containing the Matplotlib figure
        if not self.canvas:
            self.canvas = FigureCanvasTkAgg(fig,
                                       master=self.canvasArea)
            self.canvas.draw()

            # placing the canvas on the Tkinter window
            self.canvas.get_tk_widget().pack()

            # creating the Matplotlib toolbar
            toolbar = NavigationToolbar2Tk(self.canvas,
                                           self.canvasArea)
            toolbar.update()
            # placing the toolbar on the Tkinter window
            self.canvas.get_tk_widget().pack()
            self.moveYAxisButton["state"] = "normal"
        else:
            self.canvas.draw()

        self.canvas.mpl_connect('button_press_event', self.buttonPressEvent)
        self.canvas.mpl_connect('motion_notify_event', self.mouseMove)
        self.canvas.mpl_connect('button_release_event', self.buttonReleaseEvent)

    def buttonPressEvent(self, event):  # Modify from here
        x0 = event.xdata
        y0 = event.ydata
        index = self.graphXValues.index(round(x0))
        if self.modifyDataBySelection[0] == 1:
            if index or index == 0:
                if y0 - 100 < self.graphYValues[index][self.chartOnOffValue[self.selectedColumn][0]] < y0 + 100:
                    self.selectedIndex = index
                    self.STATE['click'] = 1
        elif self.modifyDataBySelection[1] == 1:
            if index or index == 0:
                if y0 - 100 < self.graphYValues[index][self.chartOnOffValue[self.selectedColumn][0]] < y0 + 100:
                    self.graphYValues = self.graphYValues[index:]
                    self.graphXValues = self.graphXValues[index:]
                    self.clearPlots()
                    self.createGraph()
                    self.modifyDataBySelection = [0, 0, 0]
                    self.populatePointListBox()
        elif self.modifyDataBySelection[2] == 1:
            if index or index == 0:
                if y0 - 100 < self.graphYValues[index][self.chartOnOffValue[self.selectedColumn][0]] < y0 + 100:
                    self.graphYValues = self.graphYValues[:index + 1]
                    self.graphXValues = self.graphXValues[:index + 1]
                    self.clearPlots()
                    self.createGraph()
                    self.modifyDataBySelection = [0, 0, 0]
                    self.populatePointListBox()

    def mouseMove(self, event):
        if self.STATE['click'] == 1 and self.modifyDataBySelection[0] == 1:
            self.updateYAxisByMove(event.ydata)

    def updateYAxisByMove(self, y0):
        self.graphYValues[self.selectedIndex][self.chartOnOffValue[self.selectedColumn][0]] = round(y0)
        self.clearPlots()
        self.createGraph()

    def buttonReleaseEvent(self, event):
        if self.modifyDataBySelection[0] == 1:
            self.STATE['click'] = 0
            self.modifyDataBySelection[0] = 0
            self.updateYAxisByMove(event.ydata)
            self.selectedIndex = None
            self.populatePointListBox()

    def open_file(self):
        self.filename = filedialog.askopenfilename(title="Open a File", filetype=(("xlxs files", ".*xlsx"),
                                                                             ("All Files", "*.")))
        if self.filename:
            try:
                self.wb = load_workbook(self.filename)
                self.sheetNames = self.wb.sheetnames
            except ValueError:
                print("File could not be opened")
                self.FileErrorLabel.config(text="File could not be opened")
            except FileNotFoundError:
                print("File Not Found")
                self.FileErrorLabel.config(text="File Not Found")
        if self.wb:
            self.callExcellValuesMethod([])

    def updateSelectionMethod(self):
        self.callExcellValuesMethod([self.secondWindow.selectedSheet1, self.secondWindow.selectedId])

    def callExcellValuesMethod(self, initialVariables):
        self.secondWindow = ExcellOpener(self.wb, initialVariables)
        self.master.wait_window(
            self.secondWindow.newWindow)  # Wait until the new window is closed to run the rest of the code
        if not self.secondWindow.cancelled:
            self.chartListbox.delete(0, END)
            self.pointListbox.delete(0, END)
            self.clearPlots()
            self.fillGraphValues(self.secondWindow.selectedSheet1)
            self.updateSelectionsButton["state"] = "normal"
            self.deleteDataBeforeButton["state"] = "normal"
            self.deleteDataAfterButton["state"] = "normal"
            self.saveDataButton["state"] = "normal"
            
    def clearPlots(self):
        if self.plot1:
            self.plot1.clear()
        if self.plot2 and self.showSecondaryChart.get() == 0:
            self.plot2.set_visible(False)
            self.plot2.clear()

    def fillGraphValues(self, selectedSheet):
        xValues = []
        self.graphYValues = []
        x = 0
        if self.titles == []:
            self.titles = self.wb[selectedSheet][1]
        for row in self.wb[selectedSheet].iter_rows(self.wb[selectedSheet].min_row,
                                                    self.wb[selectedSheet].max_row):
            passingRow = False
            passingValues = []
            for index, cell in enumerate(row):
                if index == 0 and str(cell.value) == self.secondWindow.selectedId and not passingRow:
                    passingRow = True
                if passingRow:
                    passingValues.append(cell.value)
            if passingValues:
                self.graphYValues.append(passingValues)
            if passingRow:
                xValues.append(x)
                x = x + 1
        self.populateCharListBox()
        self.graphXValues = xValues
        self.createGraph()

    def addRemoveChart(self):
        self.chartOnOffValue[self.selectedColumn][1] = not self.chartOnOffValue[self.selectedColumn][1]
        self.clearPlots()
        self.createGraph()

    def secondaryChartView(self):
        self.clearPlots()
        self.createGraph()

    def populatePointListBox(self):
        for item in self.graphYValues:
            self.pointListbox.insert(END, item[self.chartOnOffValue[self.selectedColumn][0]])

    def populateCharListBox(self):
        self.chartListbox.delete(0, END)
        for index, chartName in enumerate(self.charts):
            self.chartListbox.insert(END, chartName)
            self.chartListbox.itemconfig(END, {'fg': self.chartNameColors[index]})
        self.selectChart(0)
        self.chartOnOffValue[0][1] = True

    def selectChart(self, chartIndex):
        self.chartListbox.itemconfig(self.selectedColumn, {'bg': 'white'})
        self.chartListbox.itemconfig(chartIndex, {'bg': 'yellow'})
        self.selectedColumn = chartIndex
        self.populatePointListBox()

    def populatePointListBox(self):
        self.pointListbox.delete(0, END)
        for value in self.graphYValues:
            self.pointListbox.insert(END, value[self.chartOnOffValue[self.selectedColumn][0]])

    def saveDataMethod(self):
        self.openWindow()

    def openWindow(self):
        self.newWindow = Toplevel()
        self.newWindow.geometry("600x600")
        self.newWindow.grab_set()
        self.newWindow.protocol("WM_DELETE_WINDOW", self.disable_event)
        sheetNames = self.sheetNames
        if "Sheet1" in sheetNames:
            sheetNames.remove("Sheet1")

        # *** First Char Options ***
        firtChartOptionsFrame = Frame(self.newWindow)
        if len(sheetNames) > 0:
            saveTypeCheckBox = Checkbutton(firtChartOptionsFrame, text='Select Existing Sheet',variable=self.saveType, onvalue=1, offvalue=0, command=self.changeSaveType)
            self.savedSheetOptions = ttk.Combobox(firtChartOptionsFrame, state='disabled', textvariable=self.selectedSheetToSave)
            self.savedSheetOptions['values'] = sheetNames
            self.savedSheetOptions.current(0)
            self.selectedSheetToSave = self.savedSheetOptions.get()
            saveTypeCheckBox.pack(padx=2, pady=2, side=TOP)
            self.savedSheetOptions.pack(padx=2, pady=2, side=TOP)
        self.savedSheet = Entry(firtChartOptionsFrame, textvariable=self.typedSheetToSave)
        self.savedSheet.pack(padx=2, pady=2, side=TOP)
        saveButton = Button(firtChartOptionsFrame, text="Save", command=self.confirmSaveButtonMethod)
        cancelButton = Button(firtChartOptionsFrame, text="Cancel", command=self.cancelButtonMethod)
        saveButton.pack(padx=2, pady=2, side=TOP)
        cancelButton.pack(padx=2, pady=2, side=TOP)
        firtChartOptionsFrame.pack(side=LEFT)


    def cancelButtonMethod(self):
        self.newWindow.grab_release()
        self.newWindow.destroy()

    def confirmSaveButtonMethod(self):
        if self.saveType.get() == 1:
            sheet = self.wb[self.selectedSheetToSave]
            self.saveToSheet(self.selectedSheetToSave)
        else:
            self.wb.create_sheet(self.savedSheet.get())
            sheet = self.wb[self.savedSheet.get()]
            self.wb.save(self.filename)
            self.saveToSheet(self.savedSheet.get())
        sheet.cell(2, 3).value = 5  # (Row, Column)

    def saveToSheet(self, sheetName):
        selectedSheet = self.wb[sheetName]
        selectedRow = 1
        deletedRows = []
        deletedRowCount = 0
        yIndex = 0
        if sheetName in self.sheetNames:
            askYesNo = askyesno(title="confirmSave",
                                message="The sheet you selected already exists. Are you sure you want to overwrite?")
            if askYesNo:
                for row in selectedSheet.iter_rows(selectedSheet.min_row,
                                                            selectedSheet.max_row):
                    if str(row[0].value) == self.secondWindow.selectedId:
                        deletedRows.append(selectedRow)
                    selectedRow = selectedRow + 1
                for thisRow in deletedRows:
                    selectedSheet.delete_rows(thisRow - deletedRowCount, 1)
                    deletedRowCount = deletedRowCount + 1
                selectedRow = 1
            else:
                return
        for index, cell in enumerate(selectedSheet[1]):
            cell.value = self.titles[index].value
        for row in selectedSheet.iter_rows(selectedSheet.min_row,
                                           selectedSheet.max_row + len(self.graphYValues)):
            if row[0].value is None:
                if yIndex < len(self.graphYValues):
                    for index, yValue in enumerate(self.graphYValues[yIndex]):
                        selectedSheet.cell(row=selectedRow, column=index + 1).value = yValue
                yIndex = yIndex + 1
            selectedRow = selectedRow + 1
        self.wb.save(self.filename)

    def changeSaveType(self):
        if self.saveType.get() == 1:
            self.savedSheetOptions["state"] = "readonly"
            self.savedSheet["state"] = "disabled"
        else:
            self.savedSheetOptions["state"] = "disabled"
            self.savedSheet["state"] = "normal"

    def doNothing(self):
        print("Doing Nothing")

    def disable_event(self):
        pass

class ExcellOpener:
    def __init__(self, excelFile, initialVariables):
        self.cancelled = False
        self.ids = None
        self.selectedSheet1 = StringVar().get()
        self.selectedId = StringVar().get()
        if len(initialVariables) > 0:
            self.selectedSheet1 = initialVariables[0]
            self.selectedId = initialVariables[1]
        self.excelFile = excelFile
        self.openWindow()

    def openWindow(self):
        self.newWindow = Toplevel()
        self.newWindow.geometry("600x600")
        self.newWindow.grab_set()
        self.newWindow.protocol("WM_DELETE_WINDOW", self.disable_event)
        sheetNames = self.excelFile.sheetnames

        # *** First Char Options ***
        firtChartOptionsFrame = Frame(self.newWindow)
        # *** Class Combobox ***
        sheetLabel = Label(firtChartOptionsFrame, text="Select Sheet")
        self.firstSheetOptions = ttk.Combobox(firtChartOptionsFrame, state='readonly', textvariable=self.selectedSheet1)
        self.firstSheetOptions['values'] = sheetNames
        self.firstSheetOptions.current(sheetNames.index("Sheet1"))
        self.selectedSheet1 = self.firstSheetOptions.get()
        self.btnFirstSheet = Button(firtChartOptionsFrame, text='Comfirm Sheet',
                                    command=self.selectSheet)
        sheetLabel.pack(side=TOP)
        self.firstSheetOptions.pack(padx=2, pady=2, side=TOP)
        self.btnFirstSheet.pack(padx=2, pady=2, side=TOP)

        # *** Get Id's ***
        self.getIds()

        idLabel = Label(firtChartOptionsFrame, text="Choose an id (Required)")
        self.idOptions = ttk.Combobox(firtChartOptionsFrame, state='readonly', textvariable=self.selectedId)
        self.idOptions['values'] = self.ids
        self.idOptions.current(0)
        self.selectedId = self.idOptions.get()
        self.btnSelectId = Button(firtChartOptionsFrame, text='Comfirm Id',
                                  command=self.selectId)
        idLabel.pack(side=TOP)
        self.idOptions.pack(padx=2, pady=2, side=TOP)
        self.btnSelectId.pack(padx=2, pady=2, side=TOP)

        firtChartOptionsFrame.pack(side=LEFT)

        bottomframe = Frame(self.newWindow)
        confirmButton = Button(bottomframe, text="Confirm Selections", command=self.confirmSelectionsMethod)
        cancelButton = Button(bottomframe, text="Cancel Selections", command=self.cancelButtonMethod)
        self.confirmErrorText = Label(bottomframe, text="")

        self.confirmErrorText.pack()
        confirmButton.pack(padx=2, pady=2)
        cancelButton.pack(padx=2, pady=2)
        bottomframe.pack(side=BOTTOM)

    def selectSheet(self):
        self.selectedSheet1 = self.firstSheetOptions.get()

    def selectId(self):
        self.selectedId = self.idOptions.get()

    def getIds(self):
        self.ids = set()
        self.ids.add("-")
        for index, col in enumerate(self.excelFile[self.selectedSheet1]['A']):
            if index != 0:
                self.ids.add(str(col.value))
        self.ids = sorted(self.ids)

    def disable_event(self):
        pass

    def confirmSelectionsMethod(self):
        if self.selectedId == "-":
            self.confirmErrorText.config(text="Make sure you selected and confirmed all the required areas", fg="red")
            return
        self.newWindow.grab_release()
        self.newWindow.destroy()

    def cancelButtonMethod(self):
        self.cancelled = True
        self.newWindow.grab_release()
        self.newWindow.destroy()


# the main Tkinter window
root = Tk()  # Creates a blank window
root.state("zoomed")
App(root)
root.update()
root.mainloop()  # Keep the GUI open until close button clicked

#  https://stackoverflow.com/questions/69368997/how-to-write-a-1d-array-into-an-existing-excel-file-starting-from-a-specific-cel
#  https://www.tutorialkart.com/matplotlib-tutorial/matplotlib-draw-multiple-graphs-on-single-plot/#:~:text=Matplotlib%20%E2%80%93%20Multiple%20Graphs%20on%20same,the%20graphs%20one%20after%20another.
#  https://stackoverflow.com/questions/14762181/adding-a-y-axis-label-to-secondary-y-axis-in-matplotlib